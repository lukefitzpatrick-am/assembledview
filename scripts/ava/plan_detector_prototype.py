"""
Prototype structure-detector for the AVA auto-populate skill (Stage 1, deterministic).
Goal: prove we can locate, WITHOUT a fixed template, the metadata block, header row,
line-item columns, flight (burst) column band + granularity, cost columns, and junk
columns across heterogeneous media-owner plans. Validated on the 5 real examples.
Ported to TypeScript as lib/ava/autopopulate/detectPlanStructure.ts.
"""
import glob, os, re, datetime
import openpyxl
from openpyxl.utils import get_column_letter

DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")
META_LABELS = re.compile(r"^(client|campaign|demo|demographic|target|timing|agency|date|version|prepared|share option|booking)", re.I)
LINEITEM_HDRS = re.compile(r"(media description|station|network|format|buy type|length|days?|daypart|placement|entitlement|site number|qms format|latitude|spot|market|size|type|address)", re.I)
COST_HDRS = re.compile(r"(rate|value|cost|total|cpm|invest|media value|market value|budget|entitlement|impact|audience|reach|frequenc|spots|potential|install|production)", re.I)

def cval(c):
    v = c.value
    if v is None: return None
    if isinstance(v, (int, float)): return v
    if isinstance(v, (datetime.datetime, datetime.date)): return v
    s = str(v).strip()
    return s if s else None

def is_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)): return v
    if isinstance(v, str) and DATE_RE.match(v):
        try: return datetime.date.fromisoformat(v[:10])
        except: return None
    return None

def pick_data_sheet(wb):
    best, bestscore = None, -1
    for ws in wb.worksheets:
        score = ws.max_row * min(ws.max_column, 140)
        if re.search(r"double check|summary|audience|r\+f|move", ws.title, re.I):
            score *= 0.3
        if score > bestscore: best, bestscore = ws, score
    return best

def scan(ws, maxr=60, maxc=140):
    grid = {}
    for r in range(1, min(ws.max_row, maxr)+1):
        for c in range(1, min(ws.max_column, maxc)+1):
            v = cval(ws.cell(row=r, column=c))
            if v is not None: grid[(r,c)] = v
    return grid

def detect(ws):
    maxc = min(ws.max_column, 140)
    grid = scan(ws)
    ncols = {}
    date_cells = {}
    junk_cols = set()
    for (r,c),v in grid.items():
        ncols[r] = ncols.get(r,0)+1
        d = is_date(v)
        if d:
            date_cells.setdefault(r, []).append((c,d))
            if d.year < 2000: junk_cols.add(c)
        if isinstance(v,str) and v.strip().lower()=="error": junk_cols.add(c)

    date_row, date_list = None, []
    for r, cells in date_cells.items():
        valid = [(c,d) for c,d in cells if d.year>=2000]
        if len(valid) > len(date_list): date_row, date_list = r, sorted(valid)
    flight_cols = [c for c,_ in date_list]
    granularity = None
    if len(date_list) >= 2:
        deltas = [ (date_list[i+1][1]-date_list[i][1]).days for i in range(len(date_list)-1) ]
        deltas = [d for d in deltas if d>0]
        if deltas:
            med = sorted(deltas)[len(deltas)//2]
            granularity = ("weekly" if med<=8 else "lunar/4-weekly" if med<=31 else "monthly+")+f" (~{med}d)"

    hdr_row, hdr_score = None, 0
    for r in range(1, min(ws.max_row,60)+1):
        toks = sum(1 for c in range(1,maxc+1) if isinstance(grid.get((r,c)),str) and LINEITEM_HDRS.search(grid[(r,c)]))
        if toks > hdr_score: hdr_row, hdr_score = r, toks

    meta = {}
    top = hdr_row or (date_row or 10)
    for r in range(1, top):
        for c in range(1, min(maxc,8)+1):
            v = grid.get((r,c))
            if isinstance(v,str) and META_LABELS.search(v):
                for cc in range(c+1, min(maxc,c+6)+1):
                    nv = grid.get((r,cc))
                    if nv is not None:
                        meta[v.strip(': ').lower()] = nv if not isinstance(nv,(datetime.date,datetime.datetime)) else str(nv)[:10]
                        break

    cost_cols = {}
    for r in ([hdr_row, (hdr_row or 0)+0] if hdr_row else []):
        for c in range(1, maxc+1):
            v = grid.get((r,c))
            if isinstance(v,str) and COST_HDRS.search(v) and c not in flight_cols and c not in junk_cols:
                cost_cols[get_column_letter(c)] = v.replace('\n',' ')[:22]

    data_rows = group_rows = 0
    first_flight_c = min(flight_cols) if flight_cols else maxc
    if hdr_row:
        for r in range(hdr_row+1, ws.max_row+1):
            left = [grid.get((r,c)) for c in range(1, min(first_flight_c,12))]
            has_desc = any(isinstance(x,str) and len(x)>1 for x in left)
            has_flight = any(isinstance(grid.get((r,c)),(int,float)) for c in flight_cols) if r<=60 else has_desc
            if has_desc and has_flight: data_rows += 1
            elif has_desc and not has_flight: group_rows += 1
    return dict(sheet=ws.title, meta=meta, header_row=hdr_row, header_tokens=hdr_score,
                date_row=date_row, flight_cols=len(flight_cols),
                flight_span=(get_column_letter(min(flight_cols))+"–"+get_column_letter(max(flight_cols))) if flight_cols else None,
                granularity=granularity, junk_cols=sorted(get_column_letter(c) for c in junk_cols)[:8],
                cost_cols=cost_cols, est_data_rows=f"~{data_rows}", est_group_rows=f"~{group_rows}")

if __name__ == "__main__":
    for f in sorted(glob.glob("*.xlsx")):
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = pick_data_sheet(wb)
        d = detect(ws)
        print("="*78)
        print(os.path.basename(f)[:52])
        print(f"  data sheet     : {d['sheet']!r}")
        print(f"  metadata       : {d['meta']}")
        print(f"  header row     : {d['header_row']} ({d['header_tokens']} line-item tokens)")
        print(f"  flight band    : {d['flight_span']}  cols={d['flight_cols']}  {d['granularity']}  (dates row {d['date_row']})")
        print(f"  junk cols      : {d['junk_cols'] or 'none'}")
        print(f"  cost cols      : {d['cost_cols']}")
        print(f"  rows           : {d['est_data_rows']} line items, {d['est_group_rows']} group/header rows")
        wb.close()
